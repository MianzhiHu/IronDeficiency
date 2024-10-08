import pandas as pd
import numpy as np
from openpyxl.reader.excel import load_workbook
from sklearn.decomposition import KernelPCA
from sklearn.model_selection import train_test_split
from sklearn.cross_decomposition import PLSRegression


# Function to check if a cell is highlighted
def is_highlighted(cell):
    if cell.fill.start_color.index == '00000000':
        return False
    return True


# extract variables of interest
def extract_variables_of_interest(sheet):
    # Iterate through the cells and find highlighted ones
    highlighted_cells = []
    reverse_cols = []
    median_cols = []

    for row in sheet.iter_rows():
        for cell in row:
            if is_highlighted(cell):
                highlighted_cells.append(cell)

    # Iterate through the cells and find the reverse-coded ones
    for row in sheet.iter_rows(min_row=2):  # Start from the second row to skip headers
        for cell in row:
            if cell.value == 'r':
                first_column_value = row[0].value  # Accessing the first column value in the same row
                second_column_value = row[1].value  # Accessing the second column value in the same row
                if first_column_value is not None:
                    reverse_cols.append(first_column_value)
                if second_column_value is not None:
                    reverse_cols.append(second_column_value)
            if cell.value == 'm':
                first_column_value = row[1].value
                if first_column_value is not None:
                    median_cols.append(first_column_value)

    return highlighted_cells, reverse_cols, median_cols


def process_workbook(file_path):
    workbook = load_workbook(file_path)
    all_highlighted_cells = []
    all_reverse_cols = []
    all_median_cols = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        highlighted_cells, reverse_cols, median_cols = extract_variables_of_interest(sheet)
        all_highlighted_cells.extend(highlighted_cells)
        all_reverse_cols.extend(reverse_cols)
        all_median_cols.extend(median_cols)

    return all_highlighted_cells, all_reverse_cols, all_median_cols


# rename the columns
def rename_columns(columns, cbcl_cols=True):
    renamed_columns = []
    for i in range(len(columns)):
        if cbcl_cols:
            if i < 55:
                renamed_columns.append(f'Q{i + 1}')
            elif i == 55:
                renamed_columns.extend([f'Q56{chr(j)}' for j in range(97, 105)])  # Q56a to Q56h
            else:
                renamed_columns.append(f'Q{i + 1}') if i + 1 < (len(columns) - 6) else None
        else:
            renamed_columns.append(f'Q{i + 1}')
    return renamed_columns


def high_corr_checker(df, cutoff=0.85, checker=False):
    # check the correlation between the questionnaire items
    corr = df.corr()

    # find the variables that are highly correlated
    high_corr = []
    for i in range(corr.shape[0]):
        for j in range(i + 1, corr.shape[0]):
            if abs(corr.iloc[i, j]) > cutoff:
                high_corr.append((corr.index[i], corr.columns[j]))

    if checker:
        if high_corr:
            print('Highly correlated variables after processing:')
            for var in high_corr:
                print(var)
        else:
            print('No more highly correlated variables after dropping')

    else:
        # print the highly correlated variables
        if high_corr:
            print('Highly correlated variables:')
            for var in high_corr:
                print(var)

        # take the mean of the highly correlated variables and drop the original variables
        for var in high_corr:
            df[var[0]] = ((df[var[0]] + df[var[1]]) / 2).round()
            df = df.drop(columns=[var[1]])

    return df
